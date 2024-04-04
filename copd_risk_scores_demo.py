# -*- coding: utf-8 -*-
"""COPD Risk Scores Demo.ipynb

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/1uXXs1Eqn49rr9ZlVzsYc4bvBGRoI8D8m
"""

import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from scipy.stats import f_oneway

def populate_excel_with_random_data(filename, num_rows, num_columns):
    # Create a new workbook
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Populate the header row
    header = ["Patient Number", "Smoking Status", "Environmental Exposures", "Occupational Exposures",
              "Genetic Predisposition", "Asthma Status", "Respiratory Infection", "FEV1 Score",
              "CVD", "Metabolic Disorders", "Muscoskeletal Disorders", "Psychological Disorders",
              "Days Since Discharge", "Treatment"]
    sheet.append(header)

    # Generate random data for each patient
    for i in range(1, num_rows):  # Skip header row, start from 2nd row
        patient_number = i
        smoking_status = np.random.randint(0, 2)
        environmental_exposures = np.random.randint(0, 2)
        occupational_exposures = np.random.randint(0, 2)
        genetic_predisposition = np.random.randint(0, 2)
        asthma_status = np.random.randint(0, 2)
        respiratory_infection = np.random.randint(0, 2)
        fev1_score = np.random.uniform(0, 100)
        cvd = np.random.randint(0, 2)
        metabolic_disorders = np.random.randint(0, 2)
        muscoskeletal_disorders = np.random.randint(0, 2)
        psychological_disorders = np.random.randint(0, 2)
        days_since_discharge = np.random.randint(0, 90)
        # Generate Treatment variable randomly
        treatment = np.random.choice(['Treatment A', 'Treatment B', 'Treatment C'])

        # Populate the row with random data
        row_data = [patient_number, smoking_status, environmental_exposures, occupational_exposures,
                    genetic_predisposition, asthma_status, respiratory_infection, fev1_score,
                    cvd, metabolic_disorders, muscoskeletal_disorders, psychological_disorders,
                    days_since_discharge, treatment]
        sheet.append(row_data)

    # Save workbook to Excel file
    wb.save(filename)
    print(f"Random data has been populated in {filename}")

def print_excel_data(filename):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    # Iterate over rows and print data
    for row in sheet.iter_rows(values_only=True):
        print(row)

if __name__ == "__main__":
    # Specify parameters
    filename = "random_data.xlsx"  # Excel file name
    num_rows = 12  # Number of rows (including header row)
    num_columns = 14  # Number of columns

    # Populate Excel spreadsheet with random data
    populate_excel_with_random_data(filename, num_rows, num_columns)

    # Print the random data
    print("Random data:")
    print_excel_data(filename)

def calculate_risk_score(smoking_status, environmental_exposures, occupational_exposures,
                         genetic_predisposition, asthma_status, respiratory_infection,
                         fev1_score, cvd, metabolic_disorders, muscoskeletal_disorders,
                         psychological_disorders, days_since_discharge):

    risk_score = 0

    risk_score += smoking_status
    risk_score += environmental_exposures
    risk_score += occupational_exposures
    risk_score += genetic_predisposition
    risk_score += asthma_status
    risk_score += respiratory_infection
    risk_score += cvd
    risk_score += metabolic_disorders
    risk_score += muscoskeletal_disorders
    risk_score += psychological_disorders

    if fev1_score <= 29:
        risk_score += 3  # Very Severe
    elif fev1_score >= 80:
        risk_score += 0  # Mild
    elif 50 <= fev1_score <= 79:
        risk_score += 1  # Moderate
    elif 30 <= fev1_score <= 49:
        risk_score += 2  # Severe

    # Consider higher risk of readmission for patients discharged within 30 days
    if days_since_discharge <= 30:
        # 20% odds of being readmitted
        risk_score += np.random.choice([0, 1], p=[0.8, 0.2])

    return risk_score

def classify_risk_level(risk_score):
    if risk_score < 4:
        return "Low Risk"
    elif 4 <= risk_score <= 7:
        return "Medium Risk"
    else:
        return "High Risk"

for patient_data in patient_risk_levels:
    print(f"Patient {patient_data[0]}: Risk Level = {patient_data[1]}")

def calculate_patient_risk_scores(filename):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    # Initialize list for storing patient risk levels
    patient_risk_levels = []

    # Iterate over rows starting from the second row (excluding header)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Extract patient data from the row
        patient_number, smoking_status, environmental_exposures, occupational_exposures, \
        genetic_predisposition, asthma_status, respiratory_infection, fev1_score, \
        cvd, metabolic_disorders, muscoskeletal_disorders, psychological_disorders, \
        days_since_discharge, treatment = row

        # Calculate exacerbation risk score
        risk_score = calculate_risk_score(smoking_status, environmental_exposures,
                                          occupational_exposures, genetic_predisposition,
                                          asthma_status, respiratory_infection, fev1_score,
                                          cvd, metabolic_disorders, muscoskeletal_disorders,
                                          psychological_disorders, days_since_discharge)

        # Classify risk level
        risk_level = classify_risk_level(risk_score)

        # Append patient number and risk level to the list
        patient_risk_levels.append((patient_number, risk_level))

    return patient_risk_levels

def plot_patient_risk_distribution(patient_risk_levels):
    # Create a color mapping for risk levels
    color_mapping = {
        "Low Risk": 'lightgreen',
        "Medium Risk": 'lightsalmon',
        "High Risk": 'lightcoral',
    }

    # Map risk levels to numerical scores
    risk_score_mapping = {
        "Low Risk": 1,
        "Medium Risk": 2,
        "High Risk": 3,
    }

    # Create dictionaries to store patient data by risk level
    risk_levels_data = {level: [] for level in color_mapping}

    # Organize patient data by risk level
    for patient_number, risk_level in patient_risk_levels:
        risk_levels_data[risk_level].append(int(patient_number))

    # Create a list of all patients (1 to 11)
    all_patients = list(range(1, 12))

    # Create empty lists to store risk levels, risk scores, and colors
    risk_levels = []
    risk_scores = []
    colors = []

    # Iterate through all patients and assign risk levels, risk scores, and colors
    for patient in all_patients:
        patient_str = str(patient)
        found = False
        for risk_level, patients in risk_levels_data.items():
            if patient in patients:
                risk_levels.append(risk_level)
                risk_scores.append(risk_score_mapping[risk_level])
                colors.append(color_mapping[risk_level])
                found = True
                break
        if not found:
            # If patient has no assigned risk level, set to "Low Risk" and assign green color
            risk_levels.append("Low Risk")
            risk_scores.append(risk_score_mapping["Low Risk"])
            colors.append(color_mapping["Low Risk"])

    # Plot patient distribution by risk level
    plt.figure(figsize=(10, 6))
    plt.bar(all_patients, risk_scores, color=colors)

    # Customize plot appearance
    plt.xlabel('Patient Number')
    plt.ylabel('Risk Score')
    plt.title('Patient Risk Scores')
    plt.xticks(all_patients)
    plt.yticks(range(1, 4), risk_score_mapping.keys())  # Adjust y-axis ticks to match the number of categories
    plt.tight_layout()

    # Create custom legend
    legend_handles = [plt.Rectangle((0,0),1,1, color=color_mapping[level], label=level) for level in risk_levels_data.keys()]
    plt.legend(handles=legend_handles, title='Risk Level', loc='upper left')

    plt.show()


if __name__ == "__main__":
    # Specify the filename
    filename = "random_data.xlsx"

    # Calculate risk levels for each patient
    patient_risk_levels = calculate_patient_risk_scores(filename)

    # Plot the patient risk distribution
    plot_patient_risk_distribution(patient_risk_levels)

import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from scipy.stats import f_oneway

# Existing functions...

import openpyxl

def calculate_patient_risk_scores(filename):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    # Initialize list for storing patient risk levels
    patient_risk_levels = []

    # Iterate over rows starting from the second row (excluding header)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Extract patient data from the row
        patient_number, smoking_status, environmental_exposures, occupational_exposures, \
        genetic_predisposition, asthma_status, respiratory_infection, fev1_score, \
        cvd, metabolic_disorders, muscoskeletal_disorders, psychological_disorders, \
        days_since_discharge, treatment = row

        # Calculate risk score
        risk_score = calculate_risk_score(smoking_status, environmental_exposures,
                                          occupational_exposures, genetic_predisposition,
                                          asthma_status, respiratory_infection, fev1_score,
                                          cvd, metabolic_disorders, muscoskeletal_disorders,
                                          psychological_disorders, days_since_discharge)

        # Classify risk level
        risk_level = classify_risk_level(risk_score)

        # Append patient number and risk level to the list
        patient_risk_levels.append((patient_number, risk_level, treatment))

    return patient_risk_levels


def plot_patient_risk_distribution(patient_risk_levels):
    # The same function as before...
    pass

if __name__ == "__main__":
    # Specify the filename
    filename = "random_data.xlsx"

    # Calculate risk levels for each patient
    patient_risk_levels = calculate_patient_risk_scores(filename)

    # Print the list of patients with their risk levels
    for patient_number, risk_level, treatment in patient_risk_levels:
        print(f"Patient {patient_number}: Risk Level = {risk_level}, Treatment = {treatment}")

    # Extract risk levels for each treatment group
    risk_levels_A = [int(patient[0]) for patient in patient_risk_levels if patient[2] == 'Treatment A']
    risk_levels_B = [int(patient[0]) for patient in patient_risk_levels if patient[2] == 'Treatment B']
    risk_levels_C = [int(patient[0]) for patient in patient_risk_levels if patient[2] == 'Treatment C']

    # Perform ANOVA to test for significant differences between treatment groups
    if risk_levels_A and risk_levels_B and risk_levels_C:  # Check if all treatment groups have at least one patient
        f_statistic, p_value = f_oneway(risk_levels_A, risk_levels_B, risk_levels_C)

        # Print the results
        print("\nANOVA Results:")
        print(f"F-statistic: {f_statistic}")
        print(f"P-value: {p_value}")
    else:
        print("Cannot perform ANOVA: One or more treatment groups have no patients.")

for patient_number, risk_level, treatment in patient_risk_levels:
    print(f"Patient {patient_number}: Risk Level = {risk_level}, Treatment = {treatment}")


risk_levels_A = [int(patient[0]) for patient in patient_risk_levels if patient[2] == 'Treatment A']
risk_levels_B = [int(patient[0]) for patient in patient_risk_levels if patient[2] == 'Treatment B']
risk_levels_C = [int(patient[0]) for patient in patient_risk_levels if patient[2] == 'Treatment C']

# Check conditions (that all treatment groups have at least one patient)
if risk_levels_A and risk_levels_B and risk_levels_C:
    f_statistic, p_value = f_oneway(risk_levels_A, risk_levels_B, risk_levels_C)

    print("\nANOVA Results:")
    print(f"F-statistic: {f_statistic}")
    print(f"P-value: {p_value}")
else:
    print("Cannot perform ANOVA: One or more treatment groups have no patients.")